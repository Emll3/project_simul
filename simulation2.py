import re
import random
from rule_2 import applyRule2
from rule_3 import applyRule3
from rule_4 import applyRule4
from functools import cmp_to_key

from helper import Exponential_distribution, Normal_distribution, Bernouilli_distribution
from slot import Slot
from patient import Patient
class Simulation:
    inputFileName: str
    D: int = 6
    amountOTSlotsPerDay: int = 10
    S: int = 32 + amountOTSlotsPerDay
    slotLength: float = float(15 / 60)
    lambdaElective: float = 28.345
    meanTardiness: float = 0
    stdevTardiness: float = 2.5
    probNoShow: float = 0.02
    meanElectiveDuration: float = 15
    stdevElectiveDuration: float = 3
    lambdaUrgent: tuple[float] = (2.5, 1.25)
    probUrgentType: tuple[float] = (0.7, 0.1, 0.1, 0.05, 0.05)
    cumulativeProbUrgentType: tuple[float] = (0.7, 0.8, 0.9, 0.95, 1.0)
    meanUrgentDuration: tuple[float] = (15, 17.5, 22.5, 30, 30)
    stdevUrgentDuration: tuple[float] = (2.5, 1, 2.5, 1, 4.5)
    weightEl: float = 1.0 / 168.0
    weightUr: float = 1.0 / 9.0

    W: int
    R: int
    d: int
    s: int
    w: int
    r: int
    rule: int
    weekSchedule: list[list[Slot]]

    patients: list[Patient]
    movingAvgElectiveAppWT: list[float]
    movingAvgElectiveScanWT: list[float]
    movingAvgUrgentScanWT: list[float]
    movingAvgOT: list[float]
    avgElectiveAppWT: float
    avgElectiveScanWT: float
    avgUrgentScanWt: float
    avgOT: float
    numberOfElectivePatientsPlanned: int
    numberOfUrgentPatientsPlanned: int

    def __init__(self, filename: str, W: int, R: int, rule: int) -> None:
        self.movingAvgElectiveAppWT = list()
        self.movingAvgElectiveScanWT = list()
        self.movingAvgUrgentScanWT = list()
        self.movingAvgOT = list()

        self.rng_el_arrival = None
        self.rng_ur_arrival = None
        self.rng_tardiness = None
        self.rng_noshow = None
        self.rng_el_duration = None
        self.rng_scan_type = None
        self.rng_ur_duration = None

        self.patients = list()
        self.inputFileName = filename
        self.W = W
        self.R = R
        self.rule = rule

        self.avgElectiveAppWT = 0
        self.avgElectiveScanWT = 0
        self.avgUrgentScanWt = 0
        self.avgOT = 0
        self.numberOfElectivePatientsPlanned = 0
        self.numberOfUrgentPatientsPlanned = 0

        self.weekSchedule = []
        for d in range(self.D):
            self.weekSchedule.append([Slot() for s in range(self.S)])

        self.movingAvgElectiveAppWT = list()
        self.movingAvgElectiveScanWT = list()
        self.movingAvgUrgentScanWT = list()
        self.movingAvgOT = list()
    def schedulePatients(self) -> None:
   
        self.patients.sort(key=lambda p: (
            p.callWeek, p.callDay, p.callTime,
            0 if p.patientType == 2 else 1
        ))

        slot_queues = {1: [], 2: []}
        for w in range(self.W):
            for d in range(self.D):
                for s in range(self.S):
                    pt = self.weekSchedule[d][s].patientType
                    if pt in slot_queues:
                        slot_queues[pt].append((w, d, s, self.weekSchedule[d][s].appTime))

        ptrs = {1: 0, 2: 0}

        numberOfElectivePerWeek = 0
        numberOfElective = 0
        previousWeek = 0

        for patient in self.patients:
            i = patient.patientType
            queue = slot_queues[i]
            ptr = ptrs[i]

            if ptr >= len(queue):
                continue

            while ptr < len(queue):
                sw, sd, ss, at = queue[ptr]
                if (sw > patient.callWeek or
                (sw == patient.callWeek and sd > patient.callDay) or
                (sw == patient.callWeek and sd == patient.callDay and at > patient.callTime)):
                    break
                ptr += 1

            if ptr >= len(queue):
                ptrs[i] = ptr
                continue

            sw, sd, ss, at = queue[ptr]
            patient.scanWeek = sw
            patient.scanDay = sd
            patient.slotNr = ss
            patient.appTime = at

            if i == 1:
                if previousWeek < sw:
                    if numberOfElectivePerWeek > 0:
                        self.movingAvgElectiveAppWT[previousWeek] /= numberOfElectivePerWeek
                    numberOfElectivePerWeek = 0
                    previousWeek = sw
                wt = patient.getAppWT()
                self.movingAvgElectiveAppWT[sw] += wt
                numberOfElectivePerWeek += 1
                self.avgElectiveAppWT += wt
                numberOfElective += 1

            ptrs[i] = ptr + 1

        if numberOfElectivePerWeek > 0:
            self.movingAvgElectiveAppWT[self.W - 1] /= numberOfElectivePerWeek
        if numberOfElective > 0:
            self.avgElectiveAppWT /= numberOfElective

    def generatePatients(self) -> None:
   
        counter = 0
        for w in range(self.W):
            for d in range(self.D):
                if d < self.D - 1:
                    t = 8.0 + Exponential_distribution(self.lambdaElective, self.rng_el_arrival) * 9.0
                    while t < 17.0:
                        tardiness = Normal_distribution(self.meanTardiness, self.stdevTardiness, self.rng_tardiness) / 60.0
                        noShow    = Bernouilli_distribution(self.probNoShow, self.rng_noshow)
                        duration  = Normal_distribution(self.meanElectiveDuration, self.stdevElectiveDuration, self.rng_el_duration) / 60.0
                        self.patients.append(Patient(counter, 1, 0, w, d, t, tardiness, noShow, duration))
                        counter += 1
                        t += Exponential_distribution(self.lambdaElective, self.rng_el_arrival) / self.lambdaElective * 9.0

                lmbd    = self.lambdaUrgent[1] if d in (3, 5) else self.lambdaUrgent[0]
                endTime = 12.0                  if d in (3, 5) else 17.0
                span    = endTime - 8.0

                t = 8.0 + Exponential_distribution(lmbd, self.rng_ur_arrival) * span
                while t < endTime:
                    scanType = self.getRandomScanType()
                    duration = Normal_distribution(
                        self.meanUrgentDuration[scanType],
                        self.stdevUrgentDuration[scanType],
                        self.rng_ur_duration
                    ) / 60.0
                    self.patients.append(Patient(counter, 2, scanType, w, d, t, 0, 0, duration))
                    counter += 1
                    t += Exponential_distribution(lmbd, self.rng_ur_arrival) * span

    def getRandomScanType(self) -> int:
        r = self.rng_scan_type.random()
        for idx, prob in enumerate(self.cumulativeProbUrgentType):
            if r < prob:
                return idx

    def getNextSlotNrFromTime(self, day: int, patientType: int, time: float) -> int:
        for s in range(self.S):
            if ((self.weekSchedule[day][s].appTime > time) and (patientType == self.weekSchedule[day][s].patientType)):
                return s
        print(f"NO SLOT EXISTS DURING TIME {time} \n")
        exit(0)

    @staticmethod
    def sortPatientsOnAppTime(patient1: Patient, patient2: Patient) -> None:
        if ((patient1.scanWeek == -1) and (patient2.scanWeek == -1)):
            if (patient1.callWeek < patient2.callWeek):
                return -1
            if (patient1.callWeek > patient2.callWeek):
                return 1
            if (patient1.callDay < patient2.callDay):
                return -1
            if (patient1.callDay > patient2.callDay):
                return 1
            if (patient1.callTime < patient2.callTime):
                return -1
            if (patient1.callTime > patient2.callTime):
                return 1
            if (patient1.patientType == 2):
                return -1
            if (patient2.patientType == 2):
                return 1
            return 0
        if (patient1.scanWeek == -1):
            return 1
        if (patient2.scanWeek == -1):
            return -1
        if (patient1.scanWeek < patient2.scanWeek):
            return -1
        if (patient1.scanWeek > patient2.scanWeek):
            return 1
        if (patient1.scanDay < patient2.scanDay):
            return -1
        if (patient1.scanDay > patient2.scanDay):
            return 1
        if (patient1.appTime < patient2.appTime):
            return -1
        if (patient1.appTime > patient2.appTime):
            return 1
        if (patient1.patientType == 2):
            return -1
        if (patient2.patientType == 2):
            return 1
        if (patient1.nr < patient2.nr):
            return -1
        if (patient1.nr > patient2.nr):
            return 1
        return 0

    @staticmethod
    def sortPatients(patient1: Patient, patient2: Patient) -> int:
        if (patient1.callWeek < patient2.callWeek):
            return -1
        if (patient1.callWeek > patient2.callWeek):
            return 1
        if (patient1.callDay < patient2.callDay):
            return -1
        if (patient1.callDay > patient2.callDay):
            return 1
        if (patient1.callTime < patient2.callTime):
            return -1
        if (patient1.callTime > patient2.callTime):
            return 1
        if (patient1.patientType == 2):
            return -1
        if (patient2.patientType == 2):
            return 1
        return 0

    def schedulePatients(self) -> None:

        self.patients = sorted(self.patients, key=cmp_to_key(Simulation.sortPatients))
        week = [0, 0]
        day = [0, 0]
        slot = [0, 0]

        for s in range(self.S):
            if (self.weekSchedule[0][s].patientType == 1):
                day[0] = 0
                slot[0] = s
                break
        for s in range(self.S):
            if (self.weekSchedule[0][s].patientType == 2):
                day[1] = 0
                slot[1] = s
                break

        previousWeek = 0
        numberOfElectivePerWeek = 0
        numberOfElective = 0
        for patient in self.patients:
            i = patient.patientType - 1
            if (week[i] < self.W):
                if (patient.callWeek > week[i]):
                    week[i] = patient.callWeek
                    day[i] = 0
                    slot[i] = self.getNextSlotNrFromTime(day[i], patient.patientType, 0)
                elif ((patient.callWeek == week[i]) and (patient.callDay > day[i])):
                    day[i] = patient.callDay
                    slot[i] = self.getNextSlotNrFromTime(day[i], patient.patientType, 0)
                if ((patient.callWeek == week[i]) and (patient.callDay == day[i]) and (patient.callTime >= self.weekSchedule[day[i]][slot[i]].appTime)):
                    for s in range(self.S - 1, -1, -1):
                        if (self.weekSchedule[day[i]][s].patientType == patient.patientType):
                            slotNr = s
                            break
                    if ((patient.patientType == 2) or (patient.callTime < self.weekSchedule[day[i]][slotNr].appTime)):
                        slot[i] = self.getNextSlotNrFromTime(day[i], patient.patientType, patient.callTime)
                    else:
                        if (day[i] < self.D - 1):
                            day[i] = day[i] + 1
                        else:
                            day[i]
                            week[i] += 1
                        if (week[i] < self.W):
                            slot[i] = self.getNextSlotNrFromTime(day[i], patient.patientType, 0)
                patient.scanWeek = week[i]
                patient.scanDay = day[i]
                patient.slotNr = slot[i]
                patient.appTime = self.weekSchedule[day[i]][slot[i]].appTime

                if (patient.patientType == 1):
                    if (previousWeek < week[i]):
                        self.movingAvgElectiveAppWT[previousWeek] /= numberOfElectivePerWeek
                        numberOfElectivePerWeek = 0
                        previousWeek = week[i]
                    wt = patient.getAppWT()
                    self.movingAvgElectiveAppWT[week[i]] += wt
                    numberOfElectivePerWeek += 1
                    self.avgElectiveAppWT += wt
                    numberOfElective += 1

                found = False
                startD = day[i]
                startS = slot[i] + 1
                for w in range(week[i], self.W):
                    for d in range(startD, self.D):
                        for s in range(startS, self.S):
                            if (self.weekSchedule[d][s].patientType == patient.patientType):
                                week[i] = w
                                day[i] = d
                                slot[i] = s
                                found = True
                                break
                        if (found):
                            break
                        startS = 0
                    if (found):
                        break
                    startD = 0
                if (not found):
                    week[i] = self.W
        self.movingAvgElectiveAppWT[self.W - 1] /= numberOfElectivePerWeek
        self.avgElectiveAppWT /= numberOfElective

    def runOneSimulation(self) -> None:
        self.generatePatients()
        self.schedulePatients()
        self.patients.sort(key=lambda p: (
            (p.scanWeek if p.scanWeek != -1 else 10**9),
            (p.scanDay  if p.scanWeek != -1 else 10**9),
            (p.appTime  if p.scanWeek != -1 else 10**9),
            0 if p.patientType == 2 else 1
        ))

        prevWeek         = 0
        prevDay          = -1
        numberOfPatientsWeek = [0, 0]
        numberOfPatients     = [0, 0]
        prevScanEndTime  = 0.0
        prevIsNoShow     = False

        ws = self.weekSchedule

        for patient in self.patients:
            if patient.scanWeek == -1:
                break

            sw  = patient.scanWeek
            sd  = patient.scanDay
            sNr = patient.slotNr
            pt  = patient.patientType

            arrivalTime = patient.appTime + patient.tardiness

            if not patient.isNoShow:
                new_day = (sw != prevWeek) or (sd != prevDay)
                if new_day:
                    patient.scanTime = arrivalTime
                elif prevIsNoShow:
                    patient.scanTime = max(ws[sd][sNr].startTime, max(prevScanEndTime, arrivalTime))
                else:
                    patient.scanTime = max(prevScanEndTime, arrivalTime)

                wt = patient.getScanWT()
                if pt == 1:
                    self.movingAvgElectiveScanWT[sw] += wt
                    self.avgElectiveScanWT += wt
                else:
                    self.movingAvgUrgentScanWT[sw] += wt
                    self.avgUrgentScanWt += wt
                numberOfPatientsWeek[pt - 1] += 1
                numberOfPatients[pt - 1]     += 1

            if prevDay > -1 and prevDay != sd:
                end_limit = 12 if prevDay in (3, 5) else 17
                ot = max(0.0, prevScanEndTime - end_limit)
                self.movingAvgOT[prevWeek] += ot
                self.avgOT += ot

            if prevWeek != sw:
                if numberOfPatientsWeek[0] > 0:
                    self.movingAvgElectiveScanWT[prevWeek] /= numberOfPatientsWeek[0]
                if numberOfPatientsWeek[1] > 0:
                    self.movingAvgUrgentScanWT[prevWeek]   /= numberOfPatientsWeek[1]
                self.movingAvgOT[prevWeek] /= self.D
                numberOfPatientsWeek = [0, 0]

            if patient.isNoShow:
                prevIsNoShow = True
                if (sw != prevWeek) or (sd != prevDay):
                    prevScanEndTime = ws[sd][sNr].startTime
            else:
                prevScanEndTime = patient.scanTime + patient.duration
                prevIsNoShow    = False

            prevWeek = sw
            prevDay  = sd

        if numberOfPatientsWeek[0] > 0:
            self.movingAvgElectiveScanWT[self.W - 1] /= numberOfPatientsWeek[0]
        if numberOfPatientsWeek[1] > 0:
            self.movingAvgUrgentScanWT[self.W - 1]   /= numberOfPatientsWeek[1]
        self.movingAvgOT[self.W - 1] /= self.D

        if numberOfPatients[0] > 0:
            self.avgElectiveScanWT /= numberOfPatients[0]
        if numberOfPatients[1] > 0:
            self.avgUrgentScanWt   /= numberOfPatients[1]
        self.avgOT /= (self.D * self.W)
    
    def setWeekSchedule(self) -> None:
        with open(self.inputFileName, 'r', encoding='utf-8-sig') as r:
            slotTypes = list(map(lambda x: re.findall('[0-9]', x), r.readlines()))
            assert len(slotTypes) == 32, "Error: there should be 32 slots (lines) in the file"
            for slotIdx, weekSlot in enumerate(slotTypes):
                assert len(weekSlot) == self.D, f"Error: there should be {self.D} days in the file (columns)"
                for slotDayIdx, inputInteger in enumerate(weekSlot):
                    self.weekSchedule[slotDayIdx][slotIdx].slotType = int(inputInteger)
                    self.weekSchedule[slotDayIdx][slotIdx].patientType = int(inputInteger)

        for d in range(self.D):
            for s in range(32, self.S):
                self.weekSchedule[d][s].slotType = 3
                self.weekSchedule[d][s].patientType = 2

        for d in range(self.D):
            time = 8
            for s in range(self.S):
                self.weekSchedule[d][s].startTime = time
                if (self.weekSchedule[d][s].slotType != 1):
                    self.weekSchedule[d][s].appTime = time
                else:
                    if (self.rule == 1):
                        self.weekSchedule[d][s].appTime = time
                    elif (self.rule == 2):
                        day_elective_slots = [idx for idx in range(self.S)
                                            if self.weekSchedule[d][idx].slotType == 1]
                        self.weekSchedule[d][s].appTime = applyRule2(
                            time, self.slotLength, self.weekSchedule,
                            day_elective_slots, s, d
                        )
                    elif (self.rule == 3):
                        day_elective_slots = [idx for idx in range(self.S)
                                            if self.weekSchedule[d][idx].slotType == 1]
                        self.weekSchedule[d][s].appTime = applyRule3(
                            time, self.slotLength, self.weekSchedule,
                            day_elective_slots, s, d
                        )
                    elif (self.rule == 4):
                        self.weekSchedule[d][s].appTime = applyRule4(
                            time, self.slotLength, self.stdevElectiveDuration
                        )
                time += self.slotLength
                if (time == 12):
                    time = 13

    def resetSystem(self, base_seed: int) -> None:
        self.patients = list()
        self.avgElectiveAppWT = 0.0
        self.avgElectiveScanWT = 0.0
        self.avgUrgentScanWt = 0.0
        self.avgOT = 0.0
        self.numberOfElectivePatientsPlanned = 0
        self.numberOfUrgentPatientsPlanned = 0

        self.movingAvgElectiveAppWT = []
        self.movingAvgElectiveScanWT = []
        self.movingAvgUrgentScanWT = []
        self.movingAvgOT = []
        for w in range(self.W):
            self.movingAvgElectiveAppWT.append(0.0)
            self.movingAvgElectiveScanWT.append(0.0)
            self.movingAvgUrgentScanWT.append(0.0)
            self.movingAvgOT.append(0.0)

        self.rng_el_arrival  = random.Random(base_seed + 101)
        self.rng_ur_arrival  = random.Random(base_seed + 202)
        self.rng_tardiness   = random.Random(base_seed + 303)
        self.rng_noshow      = random.Random(base_seed + 404)
        self.rng_el_duration = random.Random(base_seed + 505)
        self.rng_scan_type   = random.Random(base_seed + 606)
        self.rng_ur_duration = random.Random(base_seed + 707)

    def runSimulations(self, replication_rows: list = None, strategy: str = '', num_urgent: int = 0) -> dict:
        electiveAppWT: float = 0
        electiveScanWT: float = 0
        urgentScanWT: float = 0
        OT: float = 0
        OV: float = 0

        self.setWeekSchedule()
        print(f"\n[Strategy={strategy}, UrgentSlots={num_urgent}, Rule={self.rule}]")
        print("r \t elAppWT \t elScanWT \t urScanWT \t OT \t OV")
        for r in range(self.R):
            self.resetSystem(base_seed=r)
            self.runOneSimulation()
            ov_r = self.avgElectiveAppWT * self.weightEl + self.avgUrgentScanWt * self.weightUr
            electiveAppWT += self.avgElectiveAppWT
            electiveScanWT += self.avgElectiveScanWT
            urgentScanWT += self.avgUrgentScanWt
            OT += self.avgOT
            OV += ov_r
            if replication_rows is not None:
                replication_rows.append([
                    strategy, num_urgent, self.rule, r,
                    round(self.avgElectiveAppWT, 6),
                    round(self.avgElectiveScanWT, 6),
                    round(self.avgUrgentScanWt, 6),
                    round(self.avgOT, 6),
                    round(ov_r, 6),
                ])
            if r % 100 == 0:
                print(f"{r} \t {self.avgElectiveAppWT:.2f} \t\t {self.avgElectiveScanWT:.5f} \t {self.avgUrgentScanWt:.2f} \t\t {self.avgOT:.2f} \t {ov_r:.2f}")

        electiveAppWT /= self.R
        electiveScanWT /= self.R
        urgentScanWT /= self.R
        OT /= self.R
        OV /= self.R
        print("------------------------------------------------------------------------")
        print(f"AVG: \t {electiveAppWT:.2f} \t\t {electiveScanWT:.5f} \t {urgentScanWT:.2f} \t\t {OT:.2f} \t {OV:.2f}\n")

        return {
            'strategy': strategy, 'num_urgent': num_urgent, 'rule': self.rule,
            'elAppWT': electiveAppWT, 'elScanWT': electiveScanWT,
            'urScanWT': urgentScanWT, 'OT': OT, 'OV': OV,
        }


if __name__ == "__main__":
    import os
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    W = 100
    R = 30
    STRATEGIES = ['S1', 'S2', 'S3']
    URGENT_COUNTS = list(range(10, 21))
    RULES = [1, 2, 3, 4]
    OUTPUT_XLSX = "results.xlsx"
    SCHEDULES_DIR = "schedules"

    replication_rows = []
    summary_rows = []

    for strategy in STRATEGIES:
        for n_urgent in URGENT_COUNTS:
            schedule_file = os.path.join(SCHEDULES_DIR, f"input-{strategy}-{n_urgent}.txt")
            if not os.path.exists(schedule_file):
                print(f"Schedule file not found: {schedule_file} — skipping.")
                continue
            for rule in RULES:
                sim = Simulation(schedule_file, W, R, rule)
                result = sim.runSimulations(
                    replication_rows=replication_rows,
                    strategy=strategy,
                    num_urgent=n_urgent,
                )
                summary_rows.append(result)

    wb = Workbook()

    HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    BODY_FONT    = Font(name="Arial", size=10)
    ALT_FILL     = PatternFill("solid", fgColor="DCE6F1")
    CENTER       = Alignment(horizontal="center", vertical="center")
    LEFT         = Alignment(horizontal="left",   vertical="center")
    thin         = Side(style="thin", color="BFBFBF")
    BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, headers, col_widths):
        ws.append(headers)
        for col_idx, (cell, width) in enumerate(zip(ws[1], col_widths), start=1):
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = CENTER
            cell.border    = BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 18
        ws.freeze_panes = "A2"

    def style_body(ws, start_row=2):
        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
            fill = ALT_FILL if row_idx % 2 == 0 else None
            for cell in row:
                cell.font   = BODY_FONT
                cell.border = BORDER
                if fill:
                    cell.fill = fill
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000"
                    cell.alignment     = CENTER
                elif isinstance(cell.value, int):
                    cell.alignment = CENTER
                else:
                    cell.alignment = LEFT

    ws1 = wb.active
    ws1.title = "Replications"
    rep_headers  = ["Strategy", "Urgent Slots", "Rule", "Replication",
                    "elAppWT (h)", "elScanWT (h)", "urScanWT (h)", "OT (h)", "OV"]
    rep_widths   = [12, 14, 8, 14, 14, 14, 14, 10, 10]
    style_header(ws1, rep_headers, rep_widths)
    for row in replication_rows:
        ws1.append(row)
    style_body(ws1)

    ws2 = wb.create_sheet("Summary")
    sum_headers = ["Strategy", "Urgent Slots", "Rule",
                   "Avg elAppWT (h)", "Avg elScanWT (h)", "Avg urScanWT (h)", "Avg OT (h)", "Avg OV"]
    sum_widths  = [12, 14, 8, 16, 16, 16, 12, 10]
    style_header(ws2, sum_headers, sum_widths)
    for r in summary_rows:
        ws2.append([
            r['strategy'], r['num_urgent'], r['rule'],
            round(r['elAppWT'],  6), round(r['elScanWT'], 6),
            round(r['urScanWT'], 6), round(r['OT'],       6),
            round(r['OV'],       6),
        ])
    style_body(ws2)

    wb.save(OUTPUT_XLSX)
    print(f"\nResults written to {OUTPUT_XLSX}")
