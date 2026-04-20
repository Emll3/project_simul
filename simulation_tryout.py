"""
File created by Samuel Bakker.
Contains the Simulation-class used in simulation.py.

This is the file you should run.
Due to the "if __name__ == "__main__":"-statement.
"""
import re
import random
from functools import cmp_to_key
import math, statistics

from helper import Exponential_distribution, Normal_distribution, Bernouilli_distribution
from helper_tryout import AntitheticRNG
from slot import Slot
from patient import Patient
class Simulation:
    """
    Simulation instance

    Class Attributes
    ----------------
    inputFileName: str
    D: int
        number of days per week (NOTE: Sunday not included! so do NOT use to calculate appointment waiting time)
    amountOTSlotsPerDay: int
        number of overtime slots per day
    S: int
        number of slots per day
    slotLength: float
        duration of a slot (in hours)
    lamdaElective: float
    meanTardiness: float
    stdevTardiness: float
    probNoShow: float
    meanElectiveDuration: float
    stdevElectiveDuration: float
    lambdaUrgent: tuple[float]
    probUrgentType: tuple[float]
    cumulativeProbUrgentType: tuple[float]
    meanUrgentDuration: tuple[float]
    stdevUrgentDuration: tuple[float]
    weightEl: float
        objective weight elective appointment wait time
    weightUr: float
        objective weight urgent scan wait time

    Attributes
    ----------
    W: int
        Number of weeks
    R: int
        Numer of replications
    d: int
    s: int
    w: int
    r: int
    rule: int
        appointment scheduling rule
    weekSchedule: list[list[Slot]]
        list of cyclic slot schedule
    patients: list[Patient]
        list of patients
    movingAvgElectiveAppWT: list[float]
        moving average of elective appointment waiting times
    movingAvgElectiveScanWT: list[float]
        moving average elective scan waiting time
    movingAvgUrgentScanWT: list[float]
        moving average urgent scan waiting time
    movingAvgOT: list[float]
        moving average overtime
    avgElectiveAppWT: float
        average elective appointment waiting time
    avgElectiveScanWT: float
        average elective scan waiting time
    avgUrgentScanWt: float
        average urgent scan waiting time
    avgOT: float
        average overtime
    numberOfElectivePatientsPlanned: int
    numberOfUrgentPatientsPlanned: int
    """
    inputFileName: str
    D: int = 6
    amountOTSlotsPerDay: int = 10
    S: int = 32 + amountOTSlotsPerDay
    slotLength: float = float(15 / 60)
    lambdaElective: float = 28.345
    meanTardiness: float = 0
    stdevTardiness: float = 2.5
    probNoShow: float = 0.02
    meanElectiveDuration: float = 15
    stdevElectiveDuration: float = 3
    lambdaUrgent: tuple[float] = (2.5, 1.25)
    probUrgentType: tuple[float] = (0.7, 0.1, 0.1, 0.05, 0.05)
    cumulativeProbUrgentType: tuple[float] = (0.7, 0.8, 0.9, 0.95, 1.0)
    meanUrgentDuration: tuple[float] = (15, 17.5, 22.5, 30, 30)
    stdevUrgentDuration: tuple[float] = (2.5, 1, 2.5, 1, 4.5)
    weightEl: float = 1.0 / 168.0
    weightUr: float = 1.0 / 9.0

    W: int
    R: int
    d: int
    s: int
    w: int
    r: int
    rule: int
    weekSchedule: list[list[Slot]]

    # var within one simulation
    patients: list[Patient]
    movingAvgElectiveAppWT: list[float]
    movingAvgElectiveScanWT: list[float]
    movingAvgUrgentScanWT: list[float]
    movingAvgOT: list[float]
    avgElectiveAppWT: float
    avgElectiveScanWT: float
    avgUrgentScanWt: float
    avgOT: float
    numberOfElectivePatientsPlanned: int
    numberOfUrgentPatientsPlanned: int

    def __init__(self, filename: str, W: int, R: int, rule: int) -> None:
        """
        Constructor. Used to instantiate a simulation.

        Args:
            filename (str): Name of the inputfile to read in, containing a schedule
            W (int): Number of weeks to simulate (aka run length)
            R (int): Number of replications
            rule (int): The appointment scheduling rule to apply
        """
        self.movingAvgElectiveAppWT = list()
        self.movingAvgElectiveScanWT = list()
        self.movingAvgUrgentScanWT = list()
        self.movingAvgOT = list()


         # random number streams for CRN
        self.rng_el_arrival = None
        self.rng_ur_arrival = None
        self.rng_tardiness = None
        self.rng_noshow = None
        self.rng_el_duration = None
        self.rng_scan_type = None
        self.rng_ur_duration = None

        self.patients = list()
        self.inputFileName = filename
        self.W = W
        self.R = R
        self.rule = rule

        self.avgElectiveAppWT = 0
        self.avgElectiveScanWT = 0
        self.avgUrgentScanWt = 0
        self.avgOT = 0
        self.numberOfElectivePatientsPlanned = 0
        self.numberOfUrgentPatientsPlanned = 0

        # initialize weekSchedule. 6 days (excl. sunday)
        self.weekSchedule = []
        for d in range(self.D):
            self.weekSchedule.append([Slot() for s in range(self.S)])

        self.movingAvgElectiveAppWT = list()
        self.movingAvgElectiveScanWT = list()
        self.movingAvgUrgentScanWT = list()
        self.movingAvgOT = list()

    def generatePatients(self) -> None:
        """
        Create new patients and add them to the list of patients for the current simulation object.
        """
        arrivalTimeNext = 0.0
        counter = 0
        for w in range(self.W):
            for d in range(self.D):
                # Start off by generating elective patients
                if (d < self.D - 1):  # not on sunday
                    arrivalTimeNext = 8 + Exponential_distribution(self.lambdaElective, self.rng_el_arrival) * (17 - 8)
                    while (arrivalTimeNext < 17):
                        tardiness = Normal_distribution(self.meanTardiness, self.stdevTardiness, self.rng_tardiness) / 60
                        noShow = Bernouilli_distribution(self.probNoShow, self.rng_noshow)
                        duration = Normal_distribution(self.meanElectiveDuration, self.stdevElectiveDuration, self.rng_el_duration) / 60
                        # create a patient with all the calculated data and add it to the list for simulation:
                        # they arrive in the current week (outer loop: w) at the current day (inner loop: d)
                        # all patients have a time of arrival (arrivalTimeNext), can be late or early (tardiness),
                        # they also have a probability of not showing up at all (noShow), and a given duration for their procedure
                        self.patients.append(Patient(counter, 1, 0, w, d, arrivalTimeNext, tardiness, noShow, duration))
                        counter += 1
                        arrivalTimeNext += Exponential_distribution(self.lambdaElective, self.rng_el_arrival) * (17 - 8)

                lmbd = self.lambdaUrgent[0]
                endTime = 17
                # radiology dept is only open half a day on thursday and saturday
                # change the values if needed:
                if ((d == 3) or (d == 5)):
                    lmbd = self.lambdaUrgent[1]
                    endTime = 12
                arrivalTimeNext = 8 + Exponential_distribution(lmbd, self.rng_ur_arrival) * (endTime - 8)
                while (arrivalTimeNext < endTime):
                    noShow = 0  # Urgent patients always show up, would be silly otherwise
                    tardiness = 0  # Urgent patients are not planned and therefore cannot be late
                    scanType = self.getRandomScanType()
                    duration = Normal_distribution(
                        self.meanUrgentDuration[scanType],
                        self.stdevUrgentDuration[scanType],
                        self.rng_ur_duration
                    ) / 60
                    self.patients.append(Patient(counter, 2, scanType, w, d, arrivalTimeNext, tardiness, noShow, duration))
                    counter += 1
                    arrivalTimeNext += Exponential_distribution(lmbd, self.rng_ur_arrival) * (endTime - 8)

    def getRandomScanType(self) -> int:
        """
        Generate a random scanType for a patient.
        Used for Urgent patients, since the type of scan needed is unknown a priori

        Returns:
            int: integer corresponding to the type of scan
        """
        r = self.rng_scan_type.random()
        for idx, prob in enumerate(self.cumulativeProbUrgentType):
            if r < prob:
                return idx

    def getNextSlotNrFromTime(self, day: int, patientType: int, time: float) -> int:
        """
        Get the next available timeSlot based on the current time in the simulation

        Args:
            day (int): day of the simulation
            patientType (int): type (urgent or not)
            time (float): time of day

        Returns:
            int: next slotnumber
        """
        for s in range(self.S):
            if ((self.weekSchedule[day][s].appTime > time) and (patientType == self.weekSchedule[day][s].patientType)):
                return s
        print(f"NO SLOT EXISTS DURING TIME {time} \n")
        exit(0)

    @staticmethod
    def sortPatientsOnAppTime(patient1: Patient, patient2: Patient) -> None:
        """Sorting function

        Args:
            patient1 (Patient): left item
            patient2 (Patient): right item

        Returns:
            None: sorts list in place
        """
        # unscheduled patients:
        if ((patient1.scanWeek == -1) and (patient2.scanWeek == -1)):
            if (patient1.callWeek < patient2.callWeek):
                return -1
            if (patient1.callWeek > patient2.callWeek):
                return 1
            # if same week, look at days:
            if (patient1.callDay < patient2.callDay):
                return -1
            if (patient1.callDay > patient2.callDay):
                return 1
            # if same day, look at time:
            if (patient1.callTime < patient2.callTime):
                return -1
            if (patient1.callTime > patient2.callTime):
                return 1
            # if the arrival time is also the same, then urgent patients (type 2) get preference
            if (patient1.patientType == 2):
                return -1
            if (patient2.patientType == 2):
                return 1
            return 0
        if (patient1.scanWeek == -1):
            # patient1 is not scheduled yet, move backwards
            return 1
        if (patient2.scanWeek == -1):
            # patient2 is not scheduled yet, move backwards
            return -1
        # scheduled patients:
        if (patient1.scanWeek < patient2.scanWeek):
            return -1
        if (patient1.scanWeek > patient2.scanWeek):
            return 1
        if (patient1.scanDay < patient2.scanDay):
            return -1
        if (patient1.scanDay > patient2.scanDay):
            return 1
        if (patient1.appTime < patient2.appTime):
            return -1
        if (patient1.appTime > patient2.appTime):
            return 1
        # if arrival time is the same, look at urgency:
        if (patient1.patientType == 2):
            return -1
        if (patient2.patientType == 2):
            return 1
        # all the above is probably also largely summarized in the following two statements:
        if (patient1.nr < patient2.nr):
            return -1
        if (patient1.nr > patient2.nr):
            return 1
        return 0

    @staticmethod
    def sortPatients(patient1: Patient, patient2: Patient) -> int:
        # sort based on weeks
        if (patient1.callWeek < patient2.callWeek):
            return -1
        if (patient1.callWeek > patient2.callWeek):
            return 1
        # same week, look at days:
        if (patient1.callDay < patient2.callDay):
            return -1
        if (patient1.callDay > patient2.callDay):
            return 1
        # if same day, look at time of arrival:
        if (patient1.callTime < patient2.callTime):
            return -1
        if (patient1.callTime > patient2.callTime):
            return 1
        # if arrival time is the same, then urgent patients are prioritized
        if (patient1.patientType == 2):
            return -1
        if (patient2.patientType == 2):
            return 1
        return 0

    def schedulePatients(self) -> None:
        """
        Used in the sorting function.
        """
        self.patients = sorted(self.patients, key=cmp_to_key(Simulation.sortPatients))
        # now we look for the first available slot of every patient type
        week = [0, 0]  # week of next available slot
        day = [0, 0]  # day of next available slot
        slot = [0, 0]  # slotNr of next available slot

        # assumption: every day has atleast one slot of eacht patient type (so also day 0)
        # elective:
        for s in range(self.S):
            if (self.weekSchedule[0][s].patientType == 1):
                day[0] = 0
                slot[0] = s
                break
        # urgent
        for s in range(self.S):
            if (self.weekSchedule[0][s].patientType == 2):
                day[1] = 0
                slot[1] = s
                break

        previousWeek = 0
        numberOfElectivePerWeek = 0
        numberOfElective = 0
        for patient in self.patients:
            # i is used to index the w and d list.
            # this is because we only need to plan within a certain horizon (namely W amount of weeks)
            i = patient.patientType - 1
            if (week[i] < self.W):
                # look if week and day need to be updated:
                if (patient.callWeek > week[i]):
                    week[i] = patient.callWeek
                    day[i] = 0
                    # assume there is at least one slot of each patient type per day => this line will find first slot of this type
                    slot[i] = self.getNextSlotNrFromTime(day[i], patient.patientType, 0)
                elif ((patient.callWeek == week[i]) and (patient.callDay > day[i])):
                    day[i] = patient.callDay
                    slot[i] = self.getNextSlotNrFromTime(day[i], patient.patientType, 0)
                # get slot
                if ((patient.callWeek == week[i]) and (patient.callDay == day[i]) and (patient.callTime >= self.weekSchedule[day[i]][slot[i]].appTime)):
                    # as every day has all types of patienttype slots, we can look for the last slot of a certain type as follows:
                    for s in range(self.S - 1, -1, -1):
                        if (self.weekSchedule[day[i]][s].patientType == patient.patientType):
                            # this if-statement will always fire, because of the fact that all types are present in a day
                            slotNr = s
                            break
                    if ((patient.patientType == 2) or (patient.callTime < self.weekSchedule[day[i]][slotNr].appTime)):
                        slot[i] = self.getNextSlotNrFromTime(day[i], patient.patientType, patient.callTime)
                    else:
                        if (day[i] < self.D - 1):
                            day[i] = day[i] + 1
                        else:
                            day[i]
                            week[i] += 1
                        if (week[i] < self.W):
                            slot[i] = self.getNextSlotNrFromTime(day[i], patient.patientType, 0)
                # schedule patient
                patient.scanWeek = week[i]
                patient.scanDay = day[i]
                patient.slotNr = slot[i]
                patient.appTime = self.weekSchedule[day[i]][slot[i]].appTime

                if (patient.patientType == 1):
                    if (previousWeek < week[i]):
                        self.movingAvgElectiveAppWT[previousWeek] /= numberOfElectivePerWeek
                        numberOfElectivePerWeek = 0
                        previousWeek = week[i]
                    wt = patient.getAppWT()
                    self.movingAvgElectiveAppWT[week[i]] += wt
                    numberOfElectivePerWeek += 1
                    self.avgElectiveAppWT += wt
                    numberOfElective += 1

                # set next slot of the current patient type
                found = False
                startD = day[i]
                startS = slot[i] + 1
                for w in range(week[i], self.W):
                    for d in range(startD, self.D):
                        for s in range(startS, self.S):
                            if (self.weekSchedule[d][s].patientType == patient.patientType):
                                week[i] = w
                                day[i] = d
                                slot[i] = s
                                found = True
                                break
                        if (found):
                            break
                        startS = 0  # next day
                    if (found):
                        break
                    startD = 0  # next week
                if (not found):
                    week[i] = self.W
        self.movingAvgElectiveAppWT[self.W - 1] /= numberOfElectivePerWeek
        self.avgElectiveAppWT /= numberOfElective

    def runOneSimulation(self) -> None:
        """
        Run one simulation.
        1. Generatie new patients
        2. Schedule patients
        3. Sort the patients
        4. Then for every patient get the waiting time
        """
        self.generatePatients()
        self.schedulePatients()
        self.patients = sorted(self.patients, key=cmp_to_key(Simulation.sortPatientsOnAppTime))
        prevWeek = 0
        prevDay = -1
        numberOfPatientsWeek = [0, 0]
        numberOfPatients = [0, 0]
        prevScanEndTime = 0
        prevIsNoShow = False

        tard = 0  # check tardiness
        for patient in self.patients:
            if (patient.scanWeek == -1):
                # stop when a patient is not scheduled
                break

            arrivalTime = patient.appTime + patient.tardiness
            # if a patient shows up, then this ofcourse has an impact on our waiting times, so do the calc:
            if (not patient.isNoShow):
                if ((patient.scanWeek != prevWeek) or (patient.scanDay != prevDay)):
                    patient.scanTime = arrivalTime
                else:
                    if (prevIsNoShow):
                        patient.scanTime = max(self.weekSchedule[patient.scanDay][patient.slotNr].startTime, max(prevScanEndTime, arrivalTime))
                    else:
                        patient.scanTime = max(prevScanEndTime, arrivalTime)
                wt = patient.getScanWT()
                if (patient.patientType == 1):
                    # if non urgent
                    self.movingAvgElectiveScanWT[patient.scanWeek] += wt
                    self.avgElectiveScanWT += wt
                else:
                    self.movingAvgUrgentScanWT[patient.scanWeek] += wt
                    self.avgUrgentScanWt += wt
                # update count of patienttype this week
                numberOfPatientsWeek[patient.patientType - 1] += 1
                numberOfPatients[patient.patientType - 1] += 1
            # for overtime, day 3 and 5 are halfdays:
            if ((prevDay > -1) and (prevDay != patient.scanDay)):
                if ((prevDay == 3) or (prevDay == 5)):
                    self.movingAvgOT[prevWeek] += max(0, prevScanEndTime - 13)
                    self.avgOT += max(0.0, prevScanEndTime - 13)
                else:
                    self.movingAvgOT[prevWeek] += max(0, prevScanEndTime - 17)
                    self.avgOT += max(0.0, prevScanEndTime - 17)

            if (prevWeek != patient.scanWeek):
                self.movingAvgElectiveScanWT[prevWeek] /= numberOfPatientsWeek[0]
                self.movingAvgUrgentScanWT[prevWeek] /= numberOfPatientsWeek[1]
                self.movingAvgOT[prevWeek] /= self.D
                numberOfPatientsWeek[0] = 0
                numberOfPatientsWeek[1] = 0
            if (patient.isNoShow):
                # for a noshow, the prevscantime does not change bc the last time someone was scanned doesnot changed
                prevIsNoShow = True
                if ((patient.scanWeek != prevWeek) or (patient.scanDay != prevDay)):
                    prevScanEndTime = self.weekSchedule[patient.scanDay][patient.slotNr].startTime
            else:
                prevScanEndTime = patient.scanTime + patient.duration
                prevIsNoShow = False
            prevWeek = patient.scanWeek
            prevDay = patient.scanDay
            tard += patient.tardiness
        # moving averages of lastweek
        self.movingAvgElectiveScanWT[self.W - 1] /= numberOfPatientsWeek[0]
        self.movingAvgUrgentScanWT[self.W - 1] /= numberOfPatientsWeek[1]
        self.movingAvgOT[self.W - 1] /= self.D
        # calc objective values
        self.avgElectiveScanWT /= numberOfPatients[0]
        self.avgUrgentScanWt /= numberOfPatients[1]
        self.avgOT /= (self.D * self.W)

    def setWeekSchedule(self) -> None:
        """
        This method sets a cyclic slot schedule based on a given input file and applied rules.

        @Students, if you choose to use Python, make sure to implement the required rules here.
        """
        # read file:
        # NOTE we assume utf-8-sig, as many students will probably be working with MS
        with open(self.inputFileName, 'r', encoding='utf-8-sig') as r:
            slotTypes = list(map(lambda x: re.findall('[0-9]', x), r.readlines()))
            assert len(slotTypes) == 32, "Error: there should be 32 slots (lines) in the file"
            for slotIdx, weekSlot in enumerate(slotTypes):
                assert len(weekSlot) == self.D, f"Error: there should be {self.D} days in the file (columns)"
                for slotDayIdx, inputInteger in enumerate(weekSlot):
                    self.weekSchedule[slotDayIdx][slotIdx].slotType = int(inputInteger)
                    self.weekSchedule[slotDayIdx][slotIdx].patientType = int(inputInteger)

        # set type of overtime slots (3 is urgent OT)
        for d in range(self.D):
            for s in range(32, self.S):
                self.weekSchedule[d][s].slotType = 3
                self.weekSchedule[d][s].patientType = 2

        for d in range(self.D):
            time = 8
            for s in range(self.S):
                # start time slot
                self.weekSchedule[d][s].startTime = time
                # appointment time slot
                if (self.weekSchedule[d][s].slotType != 1):
                    self.weekSchedule[d][s].appTime = time
                else:
                    if (self.rule == 1):
                        # Rule 1 – Plain FCFS: appointment time = slot start time
                        self.weekSchedule[d][s].appTime = time
                    elif (self.rule == 2):
                        # Rule 2 – Bailey-Welch: first K=2 patients get slot 0 start time;
                        # all subsequent patients get slot start time minus one slot length.
                        K = 2
                        day_elective_slots = [idx for idx in range(self.S)
                                              if self.weekSchedule[d][idx].slotType == 1]
                        pos_in_day = day_elective_slots.index(s) if s in day_elective_slots else 0
                        first_slot_time = self.weekSchedule[d][day_elective_slots[0]].startTime if day_elective_slots else time
                        if pos_in_day < K:
                            self.weekSchedule[d][s].appTime = first_slot_time
                        else:
                            self.weekSchedule[d][s].appTime = time - self.slotLength
                    elif (self.rule == 3):
                        # Rule 3 – Blocking: B=2 slots per block; all patients in a block
                        # get the appointment time of the first slot in that block.
                        B = 2
                        day_elective_slots = [idx for idx in range(self.S)
                                              if self.weekSchedule[d][idx].slotType == 1]
                        pos_in_day = day_elective_slots.index(s) if s in day_elective_slots else 0
                        block_start_pos = (pos_in_day // B) * B
                        block_start_slot = day_elective_slots[block_start_pos]
                        self.weekSchedule[d][s].appTime = self.weekSchedule[d][block_start_slot].startTime
                    elif (self.rule == 4):
                        # Rule 4 – Benchmarking: appointment time = slot start time - alpha * stdev_elective
                        alpha = 0.5
                        self.weekSchedule[d][s].appTime = time - alpha * (self.stdevElectiveDuration / 60)
                time += self.slotLength
                if (time == 12):
                    # Lunchbreak, so skip ahead
                    time = 13

    def resetSystem(self, base_seed: int, antithetic: bool = False) -> None:
        self.patients = list()
        self.avgElectiveAppWT = 0.0
        self.avgElectiveScanWT = 0.0
        self.avgUrgentScanWt = 0.0
        self.avgOT = 0.0

        self.movingAvgElectiveAppWT = []
        self.movingAvgElectiveScanWT = []
        self.movingAvgUrgentScanWT = []
        self.movingAvgOT = []
        for w in range(self.W):
            self.movingAvgElectiveAppWT.append(0.0)
            self.movingAvgElectiveScanWT.append(0.0)
            self.movingAvgUrgentScanWT.append(0.0)
            self.movingAvgOT.append(0.0)

        self.rng_el_arrival  = AntitheticRNG(base_seed + 101, antithetic)
        self.rng_ur_arrival  = AntitheticRNG(base_seed + 202, antithetic)
        self.rng_tardiness   = AntitheticRNG(base_seed + 303, antithetic)
        self.rng_noshow      = AntitheticRNG(base_seed + 404, antithetic)
        self.rng_el_duration = AntitheticRNG(base_seed + 505, antithetic)
        self.rng_scan_type   = AntitheticRNG(base_seed + 606, antithetic)
        self.rng_ur_duration = AntitheticRNG(base_seed + 707, antithetic)

        # ── Peek at the first U draw from each stream (for verification) ──────
        # We create a SEPARATE peeking rng with the same seed so the simulation
        # rng state is not consumed. These values appear in the Excel output.
        def peek(seed, anti):
            r = AntitheticRNG(seed, anti)
            return round(r.random(), 6)

        self.rng_peek = {
            'U_el_arrival':  peek(base_seed + 101, antithetic),
            'U_ur_arrival':  peek(base_seed + 202, antithetic),
            'U_tardiness':   peek(base_seed + 303, antithetic),
            'U_noshow':      peek(base_seed + 404, antithetic),
            'U_el_duration': peek(base_seed + 505, antithetic),
            'U_scan_type':   peek(base_seed + 606, antithetic),
            'U_ur_duration': peek(base_seed + 707, antithetic),
        }

    def runSimulations(self, replication_rows=None, strategy='', num_urgent=0) -> dict:

        self.setWeekSchedule()
        print(f"\n[Strategy={strategy}, UrgentSlots={num_urgent}, Rule={self.rule}]")
        print("pair  elAppWT(n)  elAppWT(a)  urScanWT(n)  urScanWT(a)  OV(n)   OV(a)   OV(pair)")

        num_pairs = self.R // 2

        # store per-pair values for variance/CI calculation
        pair_elAppWTs  = []
        pair_elScanWTs = []
        pair_urScanWTs = []
        pair_OTs       = []
        pair_OVs       = []

        # store normal and antithetic separately for correlation check
        normal_OVs = []
        anti_OVs   = []

        for r in range(num_pairs):
            # ── Normal run ───────────────────────────────────────────────────
            self.resetSystem(base_seed=r, antithetic=False)
            peek_normal = dict(self.rng_peek)
            self.runOneSimulation()
            n_elAppWT  = self.avgElectiveAppWT
            n_elScanWT = self.avgElectiveScanWT
            n_urScanWT = self.avgUrgentScanWt
            n_OT       = self.avgOT
            n_OV       = n_elAppWT * self.weightEl + n_urScanWT * self.weightUr

            # ── Antithetic run (same seed, flipped U) ────────────────────────
            self.resetSystem(base_seed=r, antithetic=True)
            peek_anti = dict(self.rng_peek)
            self.runOneSimulation()
            a_elAppWT  = self.avgElectiveAppWT
            a_elScanWT = self.avgElectiveScanWT
            a_urScanWT = self.avgUrgentScanWt
            a_OT       = self.avgOT
            a_OV       = a_elAppWT * self.weightEl + a_urScanWT * self.weightUr

            # ── Pair average (the actual AV observation) ─────────────────────
            p_elAppWT  = (n_elAppWT  + a_elAppWT)  / 2
            p_elScanWT = (n_elScanWT + a_elScanWT) / 2
            p_urScanWT = (n_urScanWT + a_urScanWT) / 2
            p_OT       = (n_OT       + a_OT)       / 2
            p_OV       = p_elAppWT * self.weightEl + p_urScanWT * self.weightUr

            pair_elAppWTs.append(p_elAppWT)
            pair_elScanWTs.append(p_elScanWT)
            pair_urScanWTs.append(p_urScanWT)
            pair_OTs.append(p_OT)
            pair_OVs.append(p_OV)
            normal_OVs.append(n_OV)
            anti_OVs.append(a_OV)

            if replication_rows is not None:
                replication_rows.append([
                    strategy, num_urgent, self.rule, r,
                    # normal run
                    round(n_elAppWT,  6), round(n_elScanWT, 6),
                    round(n_urScanWT, 6), round(n_OT,       6), round(n_OV, 6),
                    # antithetic run
                    round(a_elAppWT,  6), round(a_elScanWT, 6),
                    round(a_urScanWT, 6), round(a_OT,       6), round(a_OV, 6),
                    # pair average
                    round(p_elAppWT,  6), round(p_elScanWT, 6),
                    round(p_urScanWT, 6), round(p_OT,       6), round(p_OV, 6),
                    # U peeks — normal then antithetic
                    peek_normal['U_el_arrival'],  peek_anti['U_el_arrival'],
                    peek_normal['U_ur_arrival'],  peek_anti['U_ur_arrival'],
                    peek_normal['U_el_duration'], peek_anti['U_el_duration'],
                    peek_normal['U_ur_duration'], peek_anti['U_ur_duration'],
                ])

            if r % 10 == 0:
                print(f"{r:4d}  {n_elAppWT:.3f}  {a_elAppWT:.3f}  "
                    f"{n_urScanWT:.3f}  {a_urScanWT:.3f}  "
                    f"{n_OV:.4f}  {a_OV:.4f}  {p_OV:.4f}")

        # ── Summary statistics ────────────────────────────────────────────────
        def ci(data):
            """Return (mean, variance, stdev, half_width_95) for a list."""
            n    = len(data)
            mean = statistics.mean(data)
            var  = statistics.variance(data)   # sample variance
            sd   = math.sqrt(var)
            # t critical value for 95% CI — approximated for large n,
            # use 2.045 for n=30 pairs (df=29). Adjust if R changes.
            t_crit = 2.045  # df = num_pairs - 1 = 29 for R=30
            hw   = t_crit * sd / math.sqrt(n)
            return mean, var, sd, hw

        def correlation(xs, ys):
            """Pearson correlation between two lists."""
            n     = len(xs)
            mx    = statistics.mean(xs)
            my    = statistics.mean(ys)
            cov   = sum((x - mx) * (y - my) for x, y in zip(xs, ys)) / (n - 1)
            sx    = statistics.stdev(xs)
            sy    = statistics.stdev(ys)
            return cov / (sx * sy) if (sx > 0 and sy > 0) else 0.0

        avg_OV, var_OV, sd_OV, hw_OV = ci(pair_OVs)

        # variance of normal-only estimator (for comparison with AV)
        var_normal_OV = statistics.variance(normal_OVs)
        var_anti_OV   = statistics.variance(anti_OVs)
        corr_OV       = correlation(normal_OVs, anti_OVs)

        # theoretical AV variance = (var_n + var_a + 2*cov) / 4  / num_pairs
        # but we just report the empirical pair variance which captures the same thing
        variance_reduction_pct = (1 - var_OV / (var_normal_OV / 2)) * 100  # % reduction vs naive half

        print("------------------------------------------------------------------------")
        print(f"AVG OV (pair):  {avg_OV:.4f}  ±{hw_OV:.4f}  (95% CI)")
        print(f"Var(normal OV): {var_normal_OV:.6f}")
        print(f"Var(anti OV):   {var_anti_OV:.6f}")
        print(f"Var(pair OV):   {var_OV:.6f}")
        print(f"Corr(n, a):     {corr_OV:.4f}  ← should be negative for AV to help")
        print(f"Variance reduction vs plain: {variance_reduction_pct:.1f}%")

        return {
            'strategy': strategy, 'num_urgent': num_urgent, 'rule': self.rule,
            'elAppWT': statistics.mean(pair_elAppWTs),
            'elScanWT': statistics.mean(pair_elScanWTs),
            'urScanWT': statistics.mean(pair_urScanWTs),
            'OT': statistics.mean(pair_OTs),
            'OV': avg_OV,
            # CI and variance info
            'OV_var_normal':   var_normal_OV,
            'OV_var_anti':     var_anti_OV,
            'OV_var_pair':     var_OV,
            'OV_sd':           sd_OV,
            'OV_CI_hw':        hw_OV,
            'OV_CI_lo':        avg_OV - hw_OV,
            'OV_CI_hi':        avg_OV + hw_OV,
            'OV_corr_n_a':     corr_OV,
            'OV_var_reduction_pct': variance_reduction_pct,
        }


if __name__ == "__main__":
    import os
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Experiment settings ─────────────────────────────────────────────────
    W = 100           # Number of weeks per replication (run length)
    R = 30            # Number of replications (increase for final results, e.g. 1000)
    OUTPUT_XLSX = "resultstryout.xlsx"
    SCHEDULES_DIR = "schedules"

    STRATEGIES = ['S1', 'S2', 'S3']
    URGENT_COUNTS = list(range(10, 21))   # 10 to 20 urgent slots
    RULES = [1, 2, 3, 4]

    replication_rows = []   # all per-replication rows
    summary_rows = []       # averages per configuration

    for strategy in STRATEGIES:
        for n_urgent in URGENT_COUNTS:
            schedule_file = os.path.join(SCHEDULES_DIR, f"input-{strategy}-{n_urgent}.txt")
            if not os.path.exists(schedule_file):
                print(f"Schedule file not found: {schedule_file} — skipping.")
                continue
            for rule in RULES:
                sim = Simulation(schedule_file, W, R, rule)
                result = sim.runSimulations(
                    replication_rows=replication_rows,
                    strategy=strategy,
                    num_urgent=n_urgent,
                )
                summary_rows.append(result)

    # ── Write Excel ──────────────────────────────────────────────────────────────
    wb = Workbook()

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    BODY_FONT   = Font(name="Arial", size=10)
    ALT_FILL    = PatternFill("solid", fgColor="DCE6F1")
    GREEN_FILL  = PatternFill("solid", fgColor="E2EFDA")   # highlights good correlation
    RED_FILL    = PatternFill("solid", fgColor="FCE4D6")   # highlights positive correlation
    CENTER      = Alignment(horizontal="center", vertical="center")
    LEFT        = Alignment(horizontal="left",   vertical="center")
    thin        = Side(style="thin", color="BFBFBF")
    BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, headers, col_widths):
        ws.append(headers)
        for col_idx, (cell, width) in enumerate(zip(ws[1], col_widths), start=1):
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = CENTER
            cell.border    = BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 18
        ws.freeze_panes = "A2"

    def style_body(ws, start_row=2, highlight_col=None):
        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
            fill = ALT_FILL if row_idx % 2 == 0 else None
            for col_idx, cell in enumerate(row, start=1):
                cell.font   = BODY_FONT
                cell.border = BORDER
                if fill:
                    cell.fill = fill
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000"
                    cell.alignment     = CENTER
                elif isinstance(cell.value, int):
                    cell.alignment = CENTER
                else:
                    cell.alignment = LEFT
                # highlight correlation column green/red based on sign
                if highlight_col and col_idx == highlight_col:
                    if isinstance(cell.value, float):
                        cell.fill = GREEN_FILL if cell.value < 0 else RED_FILL

    # ── Sheet 1: Replications ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Replications"
    rep_headers = [
        "Strategy", "Urgent Slots", "Rule", "Pair nr",
        # normal
        "n_elAppWT", "n_elScanWT", "n_urScanWT", "n_OT", "n_OV",
        # antithetic
        "a_elAppWT", "a_elScanWT", "a_urScanWT", "a_OT", "a_OV",
        # pair average
        "p_elAppWT", "p_elScanWT", "p_urScanWT", "p_OT", "p_OV",
        # U peeks (CRN + AV verification)
        "U_el_arr(n)", "U_el_arr(a)",
        "U_ur_arr(n)", "U_ur_arr(a)",
        "U_el_dur(n)", "U_el_dur(a)",
        "U_ur_dur(n)", "U_ur_dur(a)",
    ]
    rep_widths = [10, 12, 6, 8,
                12, 12, 12, 10, 10,
                12, 12, 12, 10, 10,
                12, 12, 12, 10, 10,
                12, 12, 12, 12, 12, 12, 12, 12]
    style_header(ws1, rep_headers, rep_widths)
    for row in replication_rows:
        ws1.append(row)
    style_body(ws1)

    # add a note explaining the U columns
    ws1['A1'].comment = None  # clear if rerunning
    from openpyxl.comments import Comment
    note_text = (
        "U_el_arr(n): first uniform draw for elective arrivals in NORMAL run.\n"
        "U_el_arr(a): same seed but flipped (= 1 - U). Should sum to ~1.0.\n"
        "CRN check: for the same Pair nr, U values should be identical across "
        "different Strategy/Rule configurations."
    )
    ws1["U1"].comment = Comment(note_text, "Simulation")

    # ── Sheet 2: Summary with CI and variance ────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    sum_headers = [
        "Strategy", "Urgent Slots", "Rule",
        "Avg OV", "OV CI low (95%)", "OV CI high (95%)", "OV CI half-width",
        "Var OV (pair)", "Var OV (normal)", "Var OV (anti)",
        "Corr(n,a) OV",  "Var reduction %",
        "Avg elAppWT", "Avg elScanWT", "Avg urScanWT", "Avg OT",
    ]
    sum_widths = [10, 12, 6,
                12, 16, 16, 16,
                14, 16, 14,
                14, 16,
                14, 14, 14, 12]
    style_header(ws2, sum_headers, sum_widths)
    corr_col = sum_headers.index("Corr(n,a) OV") + 1   # 1-indexed for Excel
    for r in summary_rows:
        ws2.append([
            r['strategy'], r['num_urgent'], r['rule'],
            round(r['OV'],              6),
            round(r['OV_CI_lo'],        6),
            round(r['OV_CI_hi'],        6),
            round(r['OV_CI_hw'],        6),
            round(r['OV_var_pair'],     6),
            round(r['OV_var_normal'],   6),
            round(r['OV_var_anti'],     6),
            round(r['OV_corr_n_a'],     4),
            round(r['OV_var_reduction_pct'], 2),
            round(r['elAppWT'],         6),
            round(r['elScanWT'],        6),
            round(r['urScanWT'],        6),
            round(r['OT'],              6),
        ])
    style_body(ws2, highlight_col=corr_col)

    # ── Sheet 3: CRN verification ────────────────────────────────────────────
    # Shows that for the same pair nr, U values are identical across configs.
    # Pivot the replication rows: group by pair nr, show U_el_arr(n) for each config.
    ws3 = wb.create_sheet("CRN Verification")
    # collect unique configs and pair nrs
    from collections import defaultdict
    crn_data = defaultdict(dict)   # crn_data[pair_nr][(strategy, n_urgent, rule)] = U_el_arr(n)
    for row in replication_rows:
        strat, n_urg, rule, pair_nr = row[0], row[1], row[2], row[3]
        u_el_n = row[19]   # U_el_arr(n) column index in replication_rows
        crn_data[pair_nr][(strat, n_urg, rule)] = u_el_n

    configs = sorted({(row[0], row[1], row[2]) for row in replication_rows})
    crn_headers = ["Pair nr"] + [f"{s}-{u}-R{r}" for s, u, r in configs]
    crn_widths  = [8] + [14] * len(configs)
    style_header(ws3, crn_headers, crn_widths)

    for pair_nr in sorted(crn_data.keys()):
        row_data = [pair_nr]
        for cfg in configs:
            row_data.append(crn_data[pair_nr].get(cfg, "N/A"))
        ws3.append(row_data)
    style_body(ws3)

    # add explanation in cell A1 comment
    from openpyxl.comments import Comment
    ws3["A1"].comment = Comment(
        "CRN check: every value in a row should be IDENTICAL.\n"
        "Same pair number = same seed = same U draw across all configurations.\n"
        "Any difference here means CRN is broken.",
        "Simulation"
    )

    wb.save(OUTPUT_XLSX)
    print(f"\nResults written to {OUTPUT_XLSX}")