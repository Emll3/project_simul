import statistics
import math
from scipy.stats import t

import statistics
import math
from scipy.stats import t
from openpyxl import load_workbook

# Lees OV-waarden rechtstreeks uit Excel
wb = load_workbook("results.xlsx")
ws = wb["Replications"]

# Lees alleen de basisconfig: S1, 14 urgente slots, Rule 1
results = []
for row in ws.iter_rows(min_row=2, values_only=True):
    strategy = row[0]
    n_urgent = row[1]
    rule     = row[2]
    ov       = row[8]
    if strategy == 'S1' and n_urgent == 14 and rule == 1 and ov is not None:
        results.append(float(ov))

print(f"Aantal replicaties geladen: {len(results)}")

n0 = len(results)
x_bar = statistics.mean(results)
S = statistics.stdev(results)

t_val = t.ppf(0.975, df=n0 - 1)
half_width = t_val * S / math.sqrt(n0)
print(f"Gemiddelde OV:  {x_bar:.4f}")
print(f"Standaarddev:   {S:.4f}")
print(f"95% CI:         [{x_bar - half_width:.4f}, {x_bar + half_width:.4f}]")
print(f"Half-width:     {half_width:.4f}")

epsilon = 0.01 * x_bar  # 1% precisie want CI al zeer smal
if epsilon == 0:
    epsilon = 0.001

n = n0
for _ in range(50):
    t_val_iter = t.ppf(0.975, df=n - 1)
    n_new = math.ceil((t_val_iter * S / epsilon) ** 2)
    if n_new == n:
        break
    n = n_new
print(f"Vereist R: {n}")

